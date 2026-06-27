import os
from typing import Any, Dict, List, Optional

from fastapi import FastAPI, HTTPException
from pydantic import BaseModel, Field
from openpyxl import load_workbook

EXCEL_FILE = os.environ.get(
    "EXCEL_FILE",
    "AIRPORT_ZONE_CORD.xlsx" if os.path.exists("AIRPORT_ZONE_CORD.xlsx") else "AIRPORT_ZONE.xlsx",
)
SHEET_NAME = os.environ.get("SHEET_NAME", "AIRPORT_ZONE")
HUB = os.environ.get("HUB", "DXB").upper()
HUB_ZONE = os.environ.get("HUB_ZONE", "Z09").upper()

app = FastAPI(title="OD Calculation REST API", version="1.0.0")

AIRPORT_ZONE: Dict[str, str] = {}
STARTUP_ERROR = ""


def norm(v: Any) -> str:
    return str(v or "").strip().upper()


def load_airport_zone() -> None:
    """Load AIRPORT -> ZONE mapping from Excel."""
    global AIRPORT_ZONE, STARTUP_ERROR
    AIRPORT_ZONE = {}
    STARTUP_ERROR = ""

    if not os.path.exists(EXCEL_FILE):
        STARTUP_ERROR = f"Airport zone mapping not found. Please place {EXCEL_FILE} with the app."
        return

    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            STARTUP_ERROR = f"Sheet {SHEET_NAME} not found in {EXCEL_FILE}."
            return

        ws = wb[SHEET_NAME]
        headers = [norm(c.value) for c in ws[1]]
        for required in ["ZONE", "AIRPORT"]:
            if required not in headers:
                STARTUP_ERROR = "Excel must have at least ZONE and AIRPORT columns."
                return

        zone_idx = headers.index("ZONE")
        airport_idx = headers.index("AIRPORT")

        for row in ws.iter_rows(min_row=2, values_only=True):
            airport = norm(row[airport_idx])
            zone = norm(row[zone_idx])
            if airport and zone:
                AIRPORT_ZONE[airport] = zone

        if HUB not in AIRPORT_ZONE:
            AIRPORT_ZONE[HUB] = HUB_ZONE

    except Exception as exc:
        STARTUP_ERROR = f"Failed to load airport zone mapping: {exc}"


load_airport_zone()


class ItinerarySector(BaseModel):
    sno: int = Field(..., alias="Sno")
    date: str = Field(..., alias="date")
    origin: str
    destination: str
    travel_class: str = Field("Y", alias="class")

    class Config:
        allow_population_by_field_name = True


class ODRequest(BaseModel):
    itinerary: List[ItinerarySector]


class ODResponse(BaseModel):
    status: str
    count: int
    results: List[Dict[str, Any]]


def get_zone(airport: str) -> str:
    airport = norm(airport)
    if airport not in AIRPORT_ZONE:
        raise ValueError(f"Airport {airport} not found in AIRPORT_ZONE Excel mapping.")
    return AIRPORT_ZONE[airport]


def zone_num(z: str) -> int:
    try:
        return int(str(z).upper().replace("Z", ""))
    except Exception:
        return 0


def direction(z1: str, z2: str) -> int:
    a = zone_num(z1)
    b = zone_num(z2)
    if b > a:
        return 1
    if b < a:
        return -1
    return 0


def od_pair(o: str, d: str) -> str:
    return f"{o}{d}"


def sector_pair(s: Dict[str, Any]) -> str:
    return f"{s['origin']}{s['destination']}"


def build_sectors(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    sectors = []
    for r in sorted(rows, key=lambda x: int(x.get("sno", 0))):
        o = norm(r.get("origin"))
        d = norm(r.get("destination"))
        if not o or not d:
            continue

        oz = get_zone(o)
        dz = get_zone(d)
        sectors.append({
            "sno": r.get("sno"),
            "date": r.get("date"),
            "origin": o,
            "destination": d,
            "origin_zone": oz,
            "destination_zone": dz,
            "class": norm(r.get("class")) or "Y",
        })
    return sectors


def group_by_class(sectors: List[Dict[str, Any]]) -> List[List[Dict[str, Any]]]:
    groups = []
    current = []
    for s in sectors:
        if not current:
            current.append(s)
        elif s["class"] == current[-1]["class"]:
            current.append(s)
        else:
            groups.append(current)
            current = [s]
    if current:
        groups.append(current)
    return groups


def find_side_trip_indexes(sectors: List[Dict[str, Any]]) -> set:
    side_indexes = set()
    if len(sectors) < 5:
        return side_indexes

    trip_origin = sectors[0]["origin"]
    if trip_origin == HUB:
        return side_indexes

    hub_loops = []
    i = 0
    while i < len(sectors):
        if sectors[i]["origin"] == HUB and sectors[i]["destination"] != HUB:
            start = i
            end = None
            for j in range(i + 1, len(sectors)):
                if sectors[j]["origin"] != HUB and sectors[j]["destination"] == HUB:
                    end = j
                    break
            if end is not None:
                hub_loops.append((start, end))
                i = end + 1
            else:
                i += 1
        else:
            i += 1

    if not hub_loops:
        return side_indexes

    main_loop = max(hub_loops, key=lambda pair: pair[1] - pair[0])
    for start, end in hub_loops:
        if (start, end) == main_loop:
            continue
        for idx in range(start, end + 1):
            side_indexes.add(idx)
    return side_indexes


def build_virtual_legs(sectors: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    legs = []
    i = 0
    while i < len(sectors):
        s = sectors[i]
        if (
            i + 1 < len(sectors)
            and s["destination"] == HUB
            and sectors[i + 1]["origin"] == HUB
            and s["origin"] != sectors[i + 1]["destination"]
        ):
            n = sectors[i + 1]
            legs.append({
                "origin": s["origin"],
                "destination": n["destination"],
                "origin_zone": s["origin_zone"],
                "destination_zone": n["destination_zone"],
                "class": s["class"],
                "component_sectors": [sector_pair(s), sector_pair(n)],
            })
            i += 2
        else:
            legs.append({
                "origin": s["origin"],
                "destination": s["destination"],
                "origin_zone": s["origin_zone"],
                "destination_zone": s["destination_zone"],
                "class": s["class"],
                "component_sectors": [sector_pair(s)],
            })
            i += 1
    return legs


def classify_return(ods: List[Dict[str, Any]]) -> str:
    if len(ods) <= 1:
        return "ONE WAY"
    first = ods[0]
    last = ods[-1]
    if first["origin"] == last["destination"] and first["destination"] == last["origin"]:
        return "RETURN MIRRORED JOURNEY"
    return "RETURN NON MIRRORED JOURNEY"


def make_od_from_leg(leg: Dict[str, Any], turnaround="NO", reason="Sector retained as OD.") -> Dict[str, Any]:
    return {
        "origin": leg["origin"],
        "destination": leg["destination"],
        "origin_zone": leg["origin_zone"],
        "destination_zone": leg["destination_zone"],
        "class": leg["class"],
        "component_sectors": leg["component_sectors"],
        "turnaround": turnaround,
        "reasoning": reason,
    }


def calculate_main_ods(sectors: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not sectors:
        return []

    trip_origin = sectors[0]["origin"]
    legs = build_virtual_legs(sectors)

    # Non-hub continuous journey logic.
    if all(s["origin"] != HUB and s["destination"] != HUB for s in sectors):
        airports = [sectors[0]["origin"]] + [s["destination"] for s in sectors]
        zones = [sectors[0]["origin_zone"]] + [s["destination_zone"] for s in sectors]
        non_zero_signs = [
            direction(zones[i - 1], zones[i])
            for i in range(1, len(zones))
            if direction(zones[i - 1], zones[i]) != 0
        ]

        if not non_zero_signs:
            return [
                {
                    "origin": s["origin"],
                    "destination": s["destination"],
                    "origin_zone": s["origin_zone"],
                    "destination_zone": s["destination_zone"],
                    "class": s["class"],
                    "component_sectors": [sector_pair(s)],
                    "turnaround": "NO",
                    "reasoning": "Same-zone non-hub sectors retained sector-wise.",
                }
                for s in sectors
            ]

        if non_zero_signs[0] != non_zero_signs[-1]:
            turn_index = len(airports) // 2
            return [
                {
                    "origin": airports[0],
                    "destination": airports[turn_index],
                    "origin_zone": zones[0],
                    "destination_zone": zones[turn_index],
                    "class": sectors[0]["class"],
                    "component_sectors": [sector_pair(s) for s in sectors[:turn_index]],
                    "turnaround": "YES",
                    "reasoning": "Zone direction changed; turnaround point created OD break.",
                },
                {
                    "origin": airports[turn_index],
                    "destination": airports[-1],
                    "origin_zone": zones[turn_index],
                    "destination_zone": zones[-1],
                    "class": sectors[0]["class"],
                    "component_sectors": [sector_pair(s) for s in sectors[turn_index:]],
                    "turnaround": "YES",
                    "reasoning": "Inbound journey after turnaround point.",
                },
            ]

        return [{
            "origin": airports[0],
            "destination": airports[-1],
            "origin_zone": zones[0],
            "destination_zone": zones[-1],
            "class": sectors[0]["class"],
            "component_sectors": [sector_pair(s) for s in sectors],
            "turnaround": "NO",
            "reasoning": "Continuous non-hub direction; sectors combined into one OD.",
        }]

    # Pure repeated hub loops are retained sector-wise.
    if trip_origin == HUB:
        pure_hub_loops = False
        if len(sectors) % 2 == 0:
            pure_hub_loops = True
            for i in range(0, len(sectors), 2):
                if i + 1 >= len(sectors):
                    pure_hub_loops = False
                    break
                outbound = sectors[i]
                inbound = sectors[i + 1]
                if not (
                    outbound["origin"] == HUB
                    and outbound["destination"] != HUB
                    and inbound["origin"] != HUB
                    and inbound["destination"] == HUB
                ):
                    pure_hub_loops = False
                    break
        if pure_hub_loops:
            return [
                {
                    "origin": s["origin"],
                    "destination": s["destination"],
                    "origin_zone": s["origin_zone"],
                    "destination_zone": s["destination_zone"],
                    "class": s["class"],
                    "component_sectors": [sector_pair(s)],
                    "turnaround": "YES" if len(sectors) > 1 else "NO",
                    "reasoning": "Pure repeated hub loops retained sector based.",
                }
                for s in sectors
            ]

    # Simple outbound + inbound via hub.
    if trip_origin != HUB and len(legs) == 2:
        return [
            make_od_from_leg(
                leg,
                "YES",
                "Hub transfer eliminated; return journey split into outbound and inbound OD.",
            )
            for leg in legs
        ]

    # General zone-direction based break logic.
    ods = []
    start = legs[0]
    current_origin = start["origin"]
    current_origin_zone = start["origin_zone"]
    current_class = start["class"]
    current_components = []
    prev_sign = 0
    last_dest = start["destination"]
    last_dest_zone = start["destination_zone"]

    for idx, leg in enumerate(legs):
        sign = direction(leg["origin_zone"], leg["destination_zone"])
        if idx == 0:
            prev_sign = sign
            current_components.extend(leg["component_sectors"])
            last_dest = leg["destination"]
            last_dest_zone = leg["destination_zone"]
            continue

        if prev_sign != 0 and sign != 0 and sign != prev_sign:
            ods.append({
                "origin": current_origin,
                "destination": last_dest,
                "origin_zone": current_origin_zone,
                "destination_zone": last_dest_zone,
                "class": current_class,
                "component_sectors": current_components,
                "turnaround": "YES",
                "reasoning": "Zone direction changed; turnaround point created OD break.",
            })
            current_origin = leg["origin"]
            current_origin_zone = leg["origin_zone"]
            current_class = leg["class"]
            current_components = list(leg["component_sectors"])
        else:
            current_components.extend(leg["component_sectors"])

        if sign != 0:
            prev_sign = sign
        last_dest = leg["destination"]
        last_dest_zone = leg["destination_zone"]

    ods.append({
        "origin": current_origin,
        "destination": last_dest,
        "origin_zone": current_origin_zone,
        "destination_zone": last_dest_zone,
        "class": current_class,
        "component_sectors": current_components,
        "turnaround": "NO" if len(ods) == 0 else "YES",
        "reasoning": "Continuous zone sequence; sectors combined into one OD.",
    })
    return ods


def calculate_group(group: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    side_indexes = find_side_trip_indexes(group)
    main_sectors = []
    side_sectors = []

    for idx, s in enumerate(group):
        if idx in side_indexes:
            side_sectors.append(s)
        else:
            main_sectors.append(s)

    main_ods = calculate_main_ods(main_sectors)
    main_classification = classify_return(main_ods)
    results = []

    for od in main_ods:
        results.append({
            "class": od["class"],
            "ond": od_pair(od["origin"], od["destination"]),
            "zone_ond": od_pair(od["origin_zone"], od["destination_zone"]),
            "classification": main_classification,
            "component_sector": ", ".join(od["component_sectors"]),
        })

    for s in side_sectors:
        results.append({
            "class": s["class"],
            "ond": od_pair(s["origin"], s["destination"]),
            "zone_ond": od_pair(s["origin_zone"], s["destination_zone"]),
            "classification": "SIDE TRIP",
            "component_sector": sector_pair(s),
        })
    return results


def calculate_all(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    sectors = build_sectors(rows)
    if not sectors:
        raise ValueError("Please provide at least one valid itinerary sector.")

    groups = group_by_class(sectors)
    final = []
    sno = 1
    for group in groups:
        for r in calculate_group(group):
            final.append({
                "Sno": sno,
                "Ond": r["ond"],
                "Zone Ond": r["zone_ond"],
                "classification": r["classification"],
                "component sector": r["component_sector"],
            })
            sno += 1
    return final


@app.get("/health")
def health() -> Dict[str, Any]:
    return {
        "status": "ERROR" if STARTUP_ERROR else "OK",
        "excel_file": EXCEL_FILE,
        "sheet": SHEET_NAME,
        "airports_loaded": len(AIRPORT_ZONE),
        "message": STARTUP_ERROR or "Airport zone mapping loaded successfully.",
    }


@app.post("/calculate-ond", response_model=ODResponse)
def calculate_ond(request: ODRequest) -> Dict[str, Any]:
    if STARTUP_ERROR:
        raise HTTPException(status_code=500, detail=STARTUP_ERROR)

    try:
        rows = [
            {
                "sno": item.sno,
                "date": item.date,
                "origin": item.origin,
                "destination": item.destination,
                "class": item.travel_class,
            }
            for item in request.itinerary
        ]
        results = calculate_all(rows)
        return {"status": "SUCCESS", "count": len(results), "results": results}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"OD calculation failed: {exc}")
